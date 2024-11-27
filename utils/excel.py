from datetime import datetime
from os import getenv
from typing import List

from dotenv import load_dotenv
from openpyxl_image_loader import SheetImageLoader

import openpyxl

from db.models import Topic, Pool

load_dotenv()


def import_pool(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    sheet_name = "MAIN"
    sheet = wb[sheet_name]

    image_loader = SheetImageLoader(sheet)

    data = []
    errors = []
    try:
        for row_num in range(2, 1002):
            if sheet.cell(row=row_num, column=1).value is not None:
                import_id = f"{len(data) + 1}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                if all(
                        [
                            sheet.cell(row=row_num, column=2).value is not None,
                            # sheet.cell(row=row_num, column=3).value is not None,
                            # sheet.cell(row=row_num, column=5).value is not None,
                            sheet.cell(row=row_num, column=7).value is not None,
                            # sheet.cell(row=row_num, column=8).value is not None,
                            # sheet.cell(row=row_num, column=9).value is not None,
                            sheet.cell(row=row_num, column=10).value is not None,
                        ]
                ):
                    question_info = Pool(
                        import_id=import_id,
                        type="ege" if sheet.cell(row=row_num, column=1).value == "КИМ ЕГЭ" else "topic",
                        level=sheet.cell(row=row_num, column=2).value,
                        text=str(sheet.cell(row=row_num, column=3).value) if sheet.cell(row=row_num,
                                                                                        column=3).value else None,
                        answer=str(int(sheet.cell(row=row_num, column=5).value)) if sheet.cell(row=row_num,
                                                                                          column=5).value else None,
                        full_mark=sheet.cell(row=row_num, column=7).value,
                        is_rotate=1 if sheet.cell(row=row_num, column=8).value == "Да" else 0,
                        is_selfcheck=1 if sheet.cell(row=row_num, column=9).value == "Да" else 0,
                        tags_list=[tag.lower().replace("ё", "е").strip(',') for tag in
                                   sheet.cell(row=row_num, column=10).value.split(", ")],
                        created_at=datetime.now()
                    )
                    try:
                        q_image = image_loader.get(f"D{row_num}")
                        q_image.save(f"{getenv('ROOT_FOLDER')}/data/temp/q_{import_id}.png")
                        question_info.question_image = 1
                        # question_info['question_image'] = 1
                    except ValueError:
                        question_info.question_image = 0

                    try:
                        a_image = image_loader.get(f"F{row_num}")
                        a_image.save(f"{getenv('ROOT_FOLDER')}/data/temp/a_{import_id}.png")
                        question_info.answer_image = 1
                    except ValueError:
                        question_info.answer_image = 0

                    data.append(question_info)
                else:
                    # sheet.cell(row=row_num, column=11).value = "ERROR"
                    errors.append(row_num)
        return {'is_ok': True, 'data': data, 'comment': "Импорт вопросов завершён", 'errors': errors}
    except Exception as e:
        return {'is_ok': False, 'comment': f"Ошибка при чтении данных таблицы: {str(e)}", 'errors': errors}


def export_pool(data: List[Pool]):
    wb = openpyxl.load_workbook(f"{getenv('ROOT_FOLDER')}/data/excel_templates/chembot_pool_list.xlsx")

    sheet_name = "MAIN"
    sheet = wb[sheet_name]

    for index, q in enumerate(data):
        sheet.cell(row=index + 2, column=1, value=q.type)
        sheet.cell(row=index + 2, column=2, value=str(q.level))
        sheet.cell(row=index + 2, column=3, value=q.text)
        sheet.cell(row=index + 2, column=5, value=str(q.answer))
        sheet.cell(row=index + 2, column=7, value=str(q.full_mark))
        sheet.cell(row=index + 2, column=8, value="Да" if bool(q.is_rotate) else "Нет")
        sheet.cell(row=index + 2, column=9, value="Да" if bool(q.is_selfcheck) else "Нет")
        sheet.cell(row=index + 2, column=10, value=", ".join(q.tags_list))

    wb.save(f"{getenv('ROOT_FOLDER')}/data/temp/chembot_pool_list.xlsx")


def export_topics_list(db_data: List[Topic]):
    wb = openpyxl.load_workbook(f"{getenv('ROOT_FOLDER')}/data/excel_templates/chembot_topics_list.xlsx")

    sheet_name = "MAIN"
    sheet = wb[sheet_name]

    data = []
    for topic in db_data:
        name = topic.name
        volume = topic.volume
        for tag in topic.tags_list:
            data.append([name, tag, volume])

    for index, el in enumerate(data):
        sheet.cell(row=index + 2, column=1, value=el[0])
        sheet.cell(row=index + 2, column=2, value=el[1])
        sheet.cell(row=index + 2, column=3, value=el[2])

    wb.save(f"{getenv('ROOT_FOLDER')}/data/temp/chembot_topics_list.xlsx")


def import_topics_list(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    sheet_name = "MAIN"

    result = {}
    errors = []
    try:
        sheet = wb[sheet_name]
    except KeyError:
        return {'is_ok': False, 'comment': f"В файле отсутствует лист \"{sheet_name}\""}

    sheet.cell(row=1, column=4).value = "Статус"

    for row_num in range(2, 1002):
        topic_name = sheet.cell(row=row_num, column=1).value
        tag = sheet.cell(row=row_num, column=2).value
        volume = sheet.cell(row=row_num, column=3).value

        if all([topic_name, tag, volume]):

            topic_name = str(topic_name).replace("ё", "е")
            tag = str(tag).replace("ё", "е").lower()

            if volume not in result:
                result[volume] = {}

            if topic_name not in result[volume].keys():
                result[volume][topic_name] = []

            result[volume][topic_name].append(tag)

            sheet.cell(row=row_num, column=4).value = "OK"

        else:
            sheet.cell(row=row_num, column=4).value = "ERROR"

    filename = f"topics_import_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    print(result)
    wb.save(f"{getenv('ROOT_FOLDER')}/data/temp/{filename}")
    return {'is_ok': True, 'comment': "Импорт тем и тегов завершён", 'data': result, 'errors': errors,
            'filename': filename}

# print(import_pool(fr"D:\repos\chemistry_bot\test_pool.xlsx"))

if __name__ == '__main__':
    import_topics_list(
        filepath=r"C:\Users\Lario\Downloads\Telegram Desktop\чистовик.xlsx"
    )